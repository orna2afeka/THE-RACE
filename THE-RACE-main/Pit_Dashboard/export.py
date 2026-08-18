"""
export.py — export stored telemetry history to CSV or Excel
===========================================================
Reads from the local SQLite store (never Firebase). Filter by device, by which
metrics (columns) to include, and by time range.

The dashboard's download button produces a clean, readable **Excel workbook**
(`to_xlsx_bytes`): a formatted Data sheet, a Charts sheet of history graphs, and
a Faults sheet. The raw CSV writer (`to_csv_bytes`) is kept for machine use / the
CLI.

As a library:
    from export import to_xlsx_bytes, to_csv_bytes
    data, n = to_xlsx_bytes(start_ts=..., end_ts=..., metrics=["bms_soc_percent"])

As a CLI (format is chosen by the --out extension: .xlsx -> workbook, else CSV):
    python export.py --out race.xlsx
    python export.py --out race.csv
    python export.py --out soc.csv --metric bms_soc_percent --metric mms_temperature_C
    python export.py --out window.xlsx --start 2026-06-18T09:00 --end 2026-06-18T11:00
    python export.py --list-metrics
"""

import argparse
import csv
import io
import math
import sys
from datetime import datetime, timezone

import db
from constants import GEAR_RATIO, WHEEL_CIRCUMFERENCE_METERS, speed_kmh
from pit_config import DEVICE_ID

# Fixed identity/time columns that always lead the raw CSV.
_BASE_COLUMNS = ["rtdb_key", "device_id", "device_ts_epoch", "device_ts_iso", "ingested_ts_epoch"]

# Subsystem groups — what the dashboard offers as BMS / MMS / Temperature / GPS
# toggles, so users filter by system instead of remembering raw column names.
METRIC_GROUPS = {
    "BMS (battery)": ["bms_soc_percent", "bms_voltage_V", "bms_current_A"],
    "MMS (motor)": ["mms_rpm", "mms_power_W", "mms_temperature_C",
                    "mms_motor_temp_C", "mms_motor_ohms",
                    "mms_motor_map", "mms_motor_map_raw"],
    "Temperature": ["battery_temp_C"],
    "Motion / GPS": ["odometer_m", "calculated_lap", "lat", "lon"],
    "Laps / Energy": ["total_race_energy", "last_lap_energy", "last_lap_time_s",
                      "lap_source"],
    "Errors / Faults": ["bms_has_error", "bms_error_code", "bms_protections",
                        "mms_has_error", "mms_error_code", "mms_alerts"],
}


def metrics_for_groups(groups):
    """Expand a list of group names into their underlying columns,
    preserving EXPORT_COLUMNS order and dropping duplicates."""
    wanted = set()
    for g in groups or []:
        wanted.update(METRIC_GROUPS.get(g, []))
    return [m for m in db.EXPORT_COLUMNS if m in wanted]


def _iso(ts):
    if ts is None:
        return ""
    return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()


def _parse_time(value):
    """Accept a unix epoch (e.g. 1718700000) or an ISO-8601 string."""
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        pass
    dt = datetime.fromisoformat(value)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.timestamp()


def _resolve_metrics(metrics):
    """Validate the requested columns, defaulting to all of them."""
    if not metrics:
        return list(db.EXPORT_COLUMNS)
    unknown = [m for m in metrics if m not in db.EXPORT_COLUMNS]
    if unknown:
        raise ValueError(
            f"unknown metric(s): {', '.join(unknown)}. "
            f"Valid metrics: {', '.join(db.EXPORT_COLUMNS)}"
        )
    return list(metrics)


def write_csv(fileobj, start_ts=None, end_ts=None, metrics=None,
              device_id=DEVICE_ID, conn=None) -> int:
    """Write filtered rows to an open text file object. Returns the row count."""
    metrics = _resolve_metrics(metrics)
    own_conn = conn is None
    if own_conn:
        conn = db.get_conn()
    try:
        rows = db.fetch_samples(conn, start_ts=start_ts, end_ts=end_ts, device_id=device_id)
        header = _BASE_COLUMNS + metrics
        writer = csv.writer(fileobj)
        writer.writerow(header)
        for r in rows:
            writer.writerow(
                [r["rtdb_key"], r["device_id"], r["device_ts"], _iso(r["device_ts"]), r["ingested_ts"]]
                + [r[m] for m in metrics]
            )
        return len(rows)
    finally:
        if own_conn:
            conn.close()


def to_csv_bytes(start_ts=None, end_ts=None, metrics=None, device_id=DEVICE_ID) -> bytes:
    """Return the CSV as UTF-8 bytes — convenient for a Streamlit download button."""
    buf = io.StringIO()
    write_csv(buf, start_ts=start_ts, end_ts=end_ts, metrics=metrics, device_id=device_id)
    return buf.getvalue().encode("utf-8")


# ============================================================================
# EXCEL (.xlsx) export — clean Data sheet + history-graph Charts sheet + Faults
# ============================================================================
# Clean, human-readable columns for the workbook. Maps an output key ->
# (header, number_format, unit, chart_color_hex). `device_ts_iso` is the time
# axis; `speed_kmh` and `distance_km` are derived (see _cell_value). Order here
# is the display order on the Data sheet.
_XLSX_COLS = {
    "device_ts_iso":     ("Time (UTC)",          None,        None,  None),
    "speed_kmh":         ("Speed (km/h)",        "0.0",       "km/h", "00FFCC"),
    "mms_rpm":           ("Motor RPM",           "0",         "rpm",  "9B59B6"),
    "mms_power_W":       ("Motor Power (W)",     "0",         "W",    "00B3FF"),
    # This is the CONTROLLER's own temperature (byte 4 of the temp frame), not
    # the motor's — it was labelled "Motor Temp" before the motor actually had
    # a sensor, which would now collide with the real one two rows down.
    "mms_temperature_C": ("Controller Temp (°C)", "0.0",  "°C", "E74C3C"),
    "mms_motor_temp_C":  ("Motor Temp (°C)",      "0.0",  "°C", "FF5E5E"),
    "mms_motor_ohms":    ("Motor Sensor (Ω)",     "0.0",  "Ω",  "C39BD3"),
    # Map name is text (no chart colour); the raw value charts as a step trace.
    "mms_motor_map":     ("Power Map",            None,   None, None),
    "mms_motor_map_raw": ("Power Map (raw)",      "0",    None, "58D68D"),
    "bms_soc_percent":   ("Battery SoC (%)",     "0.0",       "%",    "F1C40F"),
    "bms_voltage_V":     ("Battery Voltage (V)", "0.00",      "V",    "2ECC71"),
    "bms_current_A":     ("Battery Current (A)", "0.00",      "A",    "E67E22"),
    "battery_temp_C":    ("Battery Temp (°C)", "0.0",    "°C", "FF9900"),
    "distance_km":       ("Distance (km)",       "0.000",     "km",   "1ABC9C"),
    "calculated_lap":    ("Lap",                 "0",         "#",    "7F8C9B"),
    "total_race_energy": ("Total Energy (Wh)",   "0.0",       "Wh",   "58D68D"),
    "last_lap_energy":   ("Last Lap Energy (Wh)", "0.0",      "Wh",   "45B39D"),
    "last_lap_time_s":   ("Last Lap Time (s)",   "0.000",     "s",    "5DADE2"),
    "lap_source":        ("Lap Trigger",         None,        None,   None),
    "lat":               ("Latitude",            "0.000000",  None,  None),
    "lon":               ("Longitude",           "0.000000",  None,  None),
}

# The six fault columns (the "Errors / Faults" group). If any are selected we
# emit a dedicated Faults sheet instead of six mostly-empty per-row columns.
_FAULT_COLUMNS = set(METRIC_GROUPS["Errors / Faults"])

# Cap the number of points a chart series references so the workbook stays
# snappy on a long race (the Data sheet still keeps every row).
_MAX_CHART_POINTS = 2000


def _data_columns(metrics):
    """Ordered output keys for the Data sheet, derived from the selected raw
    metrics. Adds derived Speed (needs mms_rpm) and Distance (needs odometer_m)."""
    m = set(metrics)
    present = {
        "speed_kmh": "mms_rpm" in m,
        "mms_rpm": "mms_rpm" in m,
        "mms_power_W": "mms_power_W" in m,
        "mms_temperature_C": "mms_temperature_C" in m,
        "mms_motor_temp_C": "mms_motor_temp_C" in m,
        "mms_motor_ohms": "mms_motor_ohms" in m,
        "mms_motor_map": "mms_motor_map" in m,
        "mms_motor_map_raw": "mms_motor_map_raw" in m,
        "bms_soc_percent": "bms_soc_percent" in m,
        "bms_voltage_V": "bms_voltage_V" in m,
        "bms_current_A": "bms_current_A" in m,
        "battery_temp_C": "battery_temp_C" in m,
        "distance_km": "odometer_m" in m,
        "calculated_lap": "calculated_lap" in m,
        "total_race_energy": "total_race_energy" in m,
        "last_lap_energy": "last_lap_energy" in m,
        "last_lap_time_s": "last_lap_time_s" in m,
        "lap_source": "lap_source" in m,
        "lat": "lat" in m,
        "lon": "lon" in m,
    }
    return ["device_ts_iso"] + [k for k in _XLSX_COLS if k != "device_ts_iso" and present.get(k)]


def _cell_value(key, r):
    """Value for one Data-sheet cell from a telemetry row."""
    if key == "device_ts_iso":
        return _iso(r["device_ts"])
    if key == "speed_kmh":
        return speed_kmh(r["mms_rpm"] or 0)   # shared formula — see drivetrain.py
    if key == "distance_km":
        return (r["odometer_m"] or 0) / 1000.0
    return r[key]


def _safe(v):
    """Neutralize CSV/Excel formula injection: a text cell that starts with
    = + - @ is prefixed with an apostrophe so Excel treats it as text."""
    if isinstance(v, str) and v and v[0] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def _style_header(ws, ncols, nrows):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    fill = PatternFill("solid", fgColor="1F3A5F")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    if nrows:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows + 1}"


def write_xlsx(fileobj_or_path, start_ts=None, end_ts=None, metrics=None,
               device_id=DEVICE_ID, conn=None) -> int:
    """Write a formatted Excel workbook (Data + Charts + Faults). Returns the
    Data row count. `fileobj_or_path` may be a path or a binary file object."""
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    from openpyxl.utils import get_column_letter

    metrics = _resolve_metrics(metrics)
    own_conn = conn is None
    if own_conn:
        conn = db.get_conn()
    try:
        rows = db.fetch_samples(conn, start_ts=start_ts, end_ts=end_ts, device_id=device_id)
        include_faults = any(m in _FAULT_COLUMNS for m in metrics)
        fault_rows = []
        if include_faults:
            for r in db.fetch_faults(conn, device_id=device_id):
                ts = r["device_ts"] or 0
                if (start_ts is None or ts >= start_ts) and (end_ts is None or ts <= end_ts):
                    fault_rows.append(r)
    finally:
        if own_conn:
            conn.close()

    nrows = len(rows)
    cols = _data_columns(metrics)

    wb = Workbook()

    # --- Data sheet ------------------------------------------------------- #
    ws = wb.active
    ws.title = "Data"
    ws.append([_XLSX_COLS[k][0] for k in cols])
    for r in rows:
        ws.append([_cell_value(k, r) for k in cols])
    # number formats + column widths
    numfmts = {i: _XLSX_COLS[k][1] for i, k in enumerate(cols, start=1) if _XLSX_COLS[k][1]}
    if numfmts:
        for row_cells in ws.iter_rows(min_row=2, max_row=nrows + 1):
            for i, nf in numfmts.items():
                row_cells[i - 1].number_format = nf
    for i, k in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(_XLSX_COLS[k][0]) + 2)
    _style_header(ws, len(cols), nrows)

    # --- Charts sheet (from a hidden, downsampled data block) ------------- #
    chart_keys = [k for k in cols
                  if k not in ("device_ts_iso", "lat", "lon") and _XLSX_COLS[k][3]]
    if nrows and chart_keys:
        step = max(1, math.ceil(nrows / _MAX_CHART_POINTS))
        sampled = rows[::step]
        m = len(sampled)

        cd = wb.create_sheet("_chartdata")
        cd.sheet_state = "hidden"
        cd.append(["Time"] + [_XLSX_COLS[k][0] for k in chart_keys])
        for r in sampled:
            cd.append([_iso(r["device_ts"])] + [_cell_value(k, r) for k in chart_keys])

        charts = wb.create_sheet("Charts")
        cats = Reference(cd, min_col=1, min_row=2, max_row=m + 1)
        for idx, k in enumerate(chart_keys):
            chart = LineChart()
            chart.title = _XLSX_COLS[k][0]
            chart.y_axis.title = _XLSX_COLS[k][2] or ""
            chart.x_axis.title = "Time (UTC)"
            chart.x_axis.delete = False
            chart.y_axis.delete = False
            chart.height = 7.5
            chart.width = 16
            chart.legend = None
            chart.x_axis.tickLblSkip = max(1, m // 8)
            chart.x_axis.tickMarkSkip = max(1, m // 8)
            data_ref = Reference(cd, min_col=2 + idx, min_row=1, max_row=m + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            series = chart.series[0]
            series.graphicalProperties = GraphicalProperties()
            series.graphicalProperties.line = LineProperties(solidFill=_XLSX_COLS[k][3], w=28000)
            series.smooth = False
            anchor = f"{'A' if idx % 2 == 0 else 'J'}{1 + (idx // 2) * 16}"
            charts.add_chart(chart, anchor)

    # --- Faults sheet ----------------------------------------------------- #
    if include_faults:
        fs = wb.create_sheet("Faults")
        fheaders = ["Time (UTC)", "BMS error code", "BMS protections",
                    "Motor error code", "Motor alerts"]
        fs.append(fheaders)
        if fault_rows:
            for r in fault_rows:
                fs.append([
                    _iso(r["device_ts"]),
                    r["bms_error_code"],
                    _safe(r["bms_protections"]),
                    r["mms_error_code"],
                    _safe(r["mms_alerts"]),
                ])
        else:
            fs.append(["No faults recorded in this window.", None, None, None, None])
        for i, h in enumerate(fheaders, start=1):
            fs.column_dimensions[get_column_letter(i)].width = max(16, len(h) + 2)
        _style_header(fs, len(fheaders), max(1, len(fault_rows)))

    wb.save(fileobj_or_path)
    return nrows


def to_xlsx_bytes(start_ts=None, end_ts=None, metrics=None, device_id=DEVICE_ID):
    """Return (xlsx_bytes, row_count) — for a Streamlit download button."""
    buf = io.BytesIO()
    n = write_xlsx(buf, start_ts=start_ts, end_ts=end_ts, metrics=metrics, device_id=device_id)
    return buf.getvalue(), n


def main(argv=None):
    p = argparse.ArgumentParser(
        description="Export stored telemetry history. Output format follows the "
                    "--out extension: .xlsx -> formatted Excel workbook (Data + "
                    "Charts + Faults), anything else -> raw CSV (stdout if omitted).")
    p.add_argument("--out", "-o", help="output path; .xlsx for Excel, else CSV (default: stdout CSV)")
    p.add_argument("--device", default=DEVICE_ID, help=f"device id (default: {DEVICE_ID})")
    p.add_argument("--metric", "-m", action="append",
                   help="metric column to include (repeatable; default: all)")
    p.add_argument("--group", "-g", action="append",
                   help=f"subsystem group to include (repeatable). Choices: {', '.join(METRIC_GROUPS)}")
    p.add_argument("--start", help="start time: unix epoch or ISO-8601 (inclusive)")
    p.add_argument("--end", help="end time: unix epoch or ISO-8601 (inclusive)")
    p.add_argument("--list-metrics", action="store_true", help="list valid metric names and exit")
    args = p.parse_args(argv)

    if args.list_metrics:
        print("metrics:\n  " + "\n  ".join(db.EXPORT_COLUMNS))
        print("groups:\n  " + "\n  ".join(METRIC_GROUPS))
        return 0

    try:
        start_ts = _parse_time(args.start)
        end_ts = _parse_time(args.end)
        # Combine explicit metrics with any requested groups; empty == all.
        selected = list(args.metric or []) + metrics_for_groups(args.group)
        metrics = _resolve_metrics(selected or None)
    except ValueError as e:
        print(f"error: {e}", file=sys.stderr)
        return 2

    if args.out:
        if args.out.lower().endswith(".xlsx"):
            n = write_xlsx(args.out, start_ts, end_ts, metrics, args.device)
        else:
            with open(args.out, "w", newline="", encoding="utf-8") as f:
                n = write_csv(f, start_ts, end_ts, metrics, args.device)
        print(f"wrote {n} row(s) to {args.out}")
    else:
        n = write_csv(sys.stdout, start_ts, end_ts, metrics, args.device)
        print(f"# {n} row(s)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
